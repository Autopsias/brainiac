"""XLSX handler — openpyxl. One `## Sheet: <name>` section per sheet, rendered
as a Markdown table (headers retained). Formula collapse (HARDENED, per the
session brief): prefer the last-computed CACHED value (``data_only=True``); a
formula with no cached value (never opened in Excel) falls back to the raw
formula text tagged `(formula, uncomputed)` rather than crashing or silently
dropping the cell."""
from __future__ import annotations

from pathlib import Path
from typing import Any

from .base import ExtractResult, Handler, density_gate
from .tables import rows_to_markdown

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover
    _HAS_OPENPYXL = False

MAX_XLSX_BYTES = 100 * 1024 * 1024
MAX_ROWS_PER_SHEET = 20_000  # cap runaway sheets rather than hang


def _open_workbooks(path: Path) -> tuple[Any, Any | None] | ExtractResult:
    try:
        size = path.stat().st_size
    except OSError:
        size = 0
    if size > MAX_XLSX_BYTES:
        return ExtractResult.quarantine("file_too_large")
    try:
        values = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    except Exception as exc:
        return ExtractResult.quarantine(
            "xlsx_extraction_error",
            warnings=[f"{type(exc).__name__}: {exc}"],
        )
    try:
        formulas = openpyxl.load_workbook(str(path), data_only=False, read_only=True)
    except Exception:
        formulas = None
    return values, formulas


def _sheet_rows(
    values_sheet: Any,
    formulas_sheet: Any | None,
    name: str,
    warnings: list[str],
) -> list[list[object]]:
    rows: list[list[object]] = []
    if formulas_sheet is not None:
        row_pairs = zip(values_sheet.iter_rows(), formulas_sheet.iter_rows())
    else:
        row_pairs = ((row, None) for row in values_sheet.iter_rows())
    for row_index, (value_row, formula_row) in enumerate(row_pairs, start=1):
        if row_index > MAX_ROWS_PER_SHEET:
            warnings.append(f"sheet {name!r} truncated at {MAX_ROWS_PER_SHEET} rows")
            break
        rows.append([
            _cell_value(cell.value, formula_row, column_index)
            for column_index, cell in enumerate(value_row)
        ])
    return rows


def _cell_value(value: object, formula_row: Any | None, column_index: int) -> object:
    if value is not None or formula_row is None or column_index >= len(formula_row):
        return value
    formula = formula_row[column_index].value
    if isinstance(formula, str) and formula.startswith("="):
        return f"{formula} (formula, uncomputed)"
    return value


def _render_workbooks(values: Any, formulas: Any | None) -> ExtractResult:
    sections: list[str] = []
    warnings: list[str] = []
    try:
        for name in values.sheetnames:
            formula_sheet = formulas[name] if formulas is not None else None
            rows = _sheet_rows(values[name], formula_sheet, name, warnings)
            sections.append(f"## Sheet: {name}\n\n" + rows_to_markdown(rows))
    except Exception as exc:
        return ExtractResult.quarantine(
            "xlsx_extraction_error",
            warnings=[f"{type(exc).__name__}: {exc}"],
        )
    finally:
        values.close()
        if formulas is not None:
            formulas.close()
    body = "\n".join(sections)
    reason = density_gate(body)
    if reason:
        return ExtractResult.quarantine(reason)
    return ExtractResult(
        markdown=body,
        warnings=warnings,
        metadata={"sheet_count": len(sections)},
    )


class XlsxHandler(Handler):
    extensions = (".xlsx",)
    dependency_name = "openpyxl"

    @classmethod
    def available(cls) -> bool:
        return _HAS_OPENPYXL

    @classmethod
    def extract(cls, path: Path) -> ExtractResult:
        if not _HAS_OPENPYXL:
            return ExtractResult.quarantine("missing_dependency:openpyxl")
        workbooks = _open_workbooks(path)
        if isinstance(workbooks, ExtractResult):
            return workbooks
        return _render_workbooks(*workbooks)
